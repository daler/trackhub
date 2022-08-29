import numpy as np
import trackhub
from collections import defaultdict
from openpyxl import load_workbook


def sheet_to_dict(workbook, sheet_name):
    """
    Read a two column sheet and make it into a dictionary with the first
    column as keys and the second as the values of a dictionary

    Parameters
    ----------

    filename : str
        Excel filename

    sheetname : str
        Sheet name to parse

    Return
    ------

    sheet_dict : dictionary
        dictionary the two column sheet

    """
    ws = workbook[sheet_name]
    sheet_dict = {ws["A"][i].value: ws["B"][i].value for i in range(ws.max_row)}
    return sheet_dict


def worksheet_iterator(ws):
    """
    Yields dictionaries of each row of a worksheet.

    Parameters
    ----------

    ws : openpyxl.Worksheet
    """
    header = list(ws.iter_rows(values_only=True, max_row=1))[0]
    for value in ws.iter_rows(values_only=True, min_row=2):
        yield dict(zip(header, value))


def prepare_track_dict(track_dictionary):
    """
    For user configuration convenience, the Excel spreadsheet allows parameters
    that are not valid for Track objects.

    Given a dictionary created from an Excel row, return a modified dict
    containing keys for only those parameters relevant to the respective track
    type.

    Parameters
    ----------

    track_dictionary : dict
        Excel row, parsed to dict

    Return
    ------

    track_dict_new : dict
        A new dictionary that removes the fields that are not allowed
    """

    # Remove keys with empty values (which can happen if there is a column in
    # the spreadsheet with values for some, but not all, tracks).
    new_dict = {}
    for k, v in track_dictionary.items():
        if v or not is_empty(v):
            new_dict[k] = v

    # The excel spreadsheet might have parameters that are not valid for tracks
    # Get the allowed parameters for this track type
    parameters = trackhub.constants.track_fields[new_dict["tracktype"]][:]

    # We also need to keep parameters that are allowed for all track types
    # because they aren't stored in the track_fields
    parameters.extend(trackhub.constants.track_fields["all"][:])

    # These parameters are trackhub specific parameters such as "name" and "tracktype"
    parameters.extend(trackhub.constants.trackhub_specific)

    # subgroups is also allowed
    parameters += ["subgroups"]

    track_dict_new = {k: v for k, v in new_dict.items() if k in parameters}

    return track_dict_new


def is_empty(value):
    """
    Returns True if value is None, NaN, or otherwise evaluates to False.

    Parameters
    ----------

    value : str
        Any string or nan or None

    Return
    ------

    bool
    """
    if not value:
        return True
    return not isinstance(value, str) and np.isnan(float(value))


def make_assembly(wb, d):
    """
    Makes an assembly hub. Requires sheets labeled "hub" and "genome".

    Parameters
    ----------

    wb : openpyxl Workbook
        This is the Excel imported by openpyxl

    d : dict
        Dictionary where all track objects are stored

    Return
    ------

    d : dict
        Dictionary where all track objects are stored
    """
    hub_dict = sheet_to_dict(workbook=wb, sheet_name="hub")

    hub = trackhub.Hub(**hub_dict)

    # Make the genome
    genome_dict = sheet_to_dict(sheet_name="genome")
    genome = trackhub.Assembly(**genome_dict)

    genomes_file = trackhub.GenomesFile()
    hub.add_genomes_file(genomes_file)

    # Create a trackDb and add it to the genome
    trackdb = trackhub.TrackDb()
    genome.add_trackdb(trackdb)

    # Add the genome to the genomes filename here:
    genomes_file.add_genome(genome)

    d["trackdb"] = trackdb
    d["hub"] = hub

    return d


def make_default_hub(wb, d):
    """
    Makes a hub. Requires sheets labeled "hub".

    Parameters
    ----------

    wb : openpyxl Workbook
        This is the Excel imported by openpyxl

    d : dict
        Dictionary where all track objects are stored

    Return
    ------

    d : dict
        Dictionary where all track objects are stored
    """
    hub_dict = sheet_to_dict(workbook=wb, sheet_name="hub")
    hub, genomes_file, genome, trackdb = trackhub.default_hub(**hub_dict)
    d["trackdb"] = trackdb
    d["hub"] = hub

    return d


def make_super_tracks(wb, d):
    """
    Makes super tracks.

    If no sheet labeled "super_config", return `d` unchanged.

    Parameters
    ----------

    wb : openpyxl Workbook
        This is the Excel imported by openpyxl

    d : dict
        Dictionary where all track objects are stored

    Return
    ------

    d : dict
        Dictionary where all track objects are stored
    """

    if "super_config" not in wb:
        return d

    for super_dict in worksheet_iterator(wb["super_config"]):
        supertrack_object = trackhub.SuperTrack(**super_dict)
        d["super"][super_dict["name"]] = supertrack_object
        d["trackdb"].add_tracks(supertrack_object)
    return d


def make_composite_tracks(wb, d):
    """
    Makes composite tracks.

    Parameters
    ----------

    wb : openpyxl Workbook
        This is the Excel imported by openpyxl

    d : dict
        Dictionary where all track objects are stored

    Return
    ------

    d : dict
        Dictionary where all track objects are stored
    """

    # check if there is a sheet called "composite_config"
    if "composite_config" not in wb:
        return d

    for composite_dict in worksheet_iterator(wb["composite_config"]):

        # Super track can be configured in spreadsheet but is not a valid
        # CompositeTrack argument
        super_track = composite_dict.pop("super", None)

        compositetrack_object = trackhub.CompositeTrack(**composite_dict)
        d["composite"][composite_dict["name"]] = compositetrack_object

        # if there is no super track add the composite directly to the trackdb
        if is_empty(super_track):
            d["trackdb"].add_tracks(compositetrack_object)

        # Otherwise add the composite to the supertrack; we assume that the
        # supertrack has already been added to the trackdb.
        else:
            d["super"][super_track].add_tracks(compositetrack_object)

    return d


def make_view_tracks(wb, d):
    """
    Makes view tracks.

    Parameters
    ----------

    wb : openpyxl Workbook
        This is the Excel imported by openpyxl

    d : dict
        Dictionary where all track objects are stored

    Return
    ------

    d : dict
        Dictionary where all track objects are stored
    """
    if "view_config" not in wb:
        return d

    for view_dict in worksheet_iterator(wb["view_config"]):

        # Composite can be configured in spreadsheet for a view, but it's not
        # a valid ViewTrack argument.
        composite_track = view_dict.pop("composite")

        view_track_object = trackhub.ViewTrack(**view_dict)

        d["view"][view_dict["view"]] = view_track_object

        # add the view to the associated composite track
        d["composite"][composite_track].add_tracks(view_track_object)

    return d


def make_aggregate_tracks(wb, d):
    """
    Makes aggregate tracks.

    Parameters
    ----------

    wb : openpyxl Workbook
        This is the Excel imported by openpyxl

    d : dict
        Dictionary where all track objects are stored

    Return
    ------

    d : dict
        Dictionary where all track objects are stored
    """

    if "aggregate_config" not in wb:
        return d

    for agg_dict in worksheet_iterator(wb["aggregate_config"]):

        # container and container_type can be configured in the spreadsheet for
        # an aggregate track, but they are not valid AggregateTrack arguments.
        container = agg_dict.pop("container", None)
        container_type = agg_dict.pop("container_type", None)

        agg_track_object = trackhub.AggregateTrack(**agg_dict)

        if is_empty(container_type):
            d["trackdb"].add_tracks(agg_track_object)
        else:
            d[container_type][container].add_tracks(agg_track_object)

        d["aggregate"][agg_dict["name"]] = agg_track_object

    return d


def add_tracks_and_make_subgroups(wb, d, sheets):
    """
    Adds the tracks to contianers. Also adds subgroups for each track and
    creates a dictionary of composite name to full set of subgroups

    Parameters
    ----------

    wb : openpyxl Workbook
        This is the Excel imported by openpyxl

    d : dict
        Dictionary where all track objects are stored

    Return
    ------

    d : dict
        Dictionary where all track objects are stored

    composite_to_subgroups_dict : dict
        Dictionary of sets. Keys are the composite track names. The values are
        dictionaries. The dictionaries are keyed by the name of the subgroup
        and the values are the possible options for that subgroup.
    """

    composite_to_subgroups_dict = defaultdict(lambda: defaultdict(set))
    for sheet in sheets:

        # We only want to work on sheets that define tracks; by definition this
        # is anything without "config" in the name and anything not named "hub"
        # or "genome".
        if sheet in ("hub", "genome") or "config" in sheet:
            continue

        for track_dict in worksheet_iterator(wb[sheet]):

            # container and container_type can be configured for a track in the
            # spreadsheet, but they are not valid Track arguments.
            container_type = track_dict.pop("container_type", None)
            container = track_dict.pop("container", None)

            # Tracks can only have subgroups if they are in these container
            # types. If they are not in these container types, then any
            # configured subgroups for the track will be ignored.
            if container_type == "composite" or container_type == "view":

                # Each track has a dictionary that keeps track of the subgroups
                # for that track (subgroup_dict). While we're at it, we
                # accumulate a set of all the values observed for each track so
                # that we can build subgroups later (in
                # composite_to_subgroups_dict).
                subgroup_dict = {}
                for k, v in track_dict.items():

                    # Columns in spreadsheet (keys in track_dict) that start
                    # with "subgroup_" are interpreted as being subgroup names.
                    if k.startswith("subgroup_") and not is_empty(v):

                        subgroup_name = k.replace("subgroup_", "")
                        subgroup_dict[subgroup_name] = v

                        if container_type == "view":

                            # Recall that subgroups can be defined for a track
                            # but that track is only listed as part of a view
                            # (not directly on a composite). So here we look up
                            # the parent composite so we know what to add the
                            # subgroup to.
                            parent_composite_name = d["view"][container].parent.name
                            composite_to_subgroups_dict[parent_composite_name][
                                subgroup_name
                            ] |= set([v])
                        else:
                            composite_to_subgroups_dict[container][
                                subgroup_name
                            ] |= set([v])

                # If there are subgroups then add them to the track
                if subgroup_dict:
                    track_dict["subgroups"] = subgroup_dict

            # Validate track items, so that only valid Track arguments remain.
            track_dict = prepare_track_dict(track_dict)
            track = trackhub.Track(**track_dict)

            if is_empty(container_type):
                d["trackdb"].add_tracks(track)
            else:
                d[container_type][container].add_tracks(track)

    return d, composite_to_subgroups_dict


def add_subgroups_to_composite(composite_to_subgroups_dict, d):
    """
    Adds the subgroups to the composite tracks

    Parameters
    ----------

    d : dict
        Dictionary where all track objects are stored

    composite_to_subgroups_dict : dict
        Dictionary of sets. Keys are the composite track names. The values are
        dictionaries. The dictionaries are keyed by the name of the subgroup
        and the values are the possible options for that subgroup. E.g.:

            {
              'composite1': {
                  'celltype': set(['A', 'B', 'C']),
                  'treatment': set(['control', 'treatment']),
              },
              ...
            }

    Return
    ------

    d : dict
        Dictionary where all track objects are stored, where composite tracks
        have had the appropriate SubGroupDefinitions added to them.
    """
    for composite, subgroup_options in composite_to_subgroups_dict.items():

        # Create a list of subgroup objects to add to the composite
        subgroups = []
        for subgroup_name, options in subgroup_options.items():
            mapping_dict = {group: group for group in options}
            s = trackhub.SubGroupDefinition(
                name=subgroup_name, label=subgroup_name, mapping=mapping_dict
            )
            subgroups.append(s)
        d["composite"][composite].add_subgroups(subgroups)

    return d


def main(filename, staging):
    wb = load_workbook(filename, data_only=True)

    # Get the sheet names from the Excel Workbook
    sheets = load_workbook(filename).sheetnames

    # This dictionary holds all of the containter tracks and hub and trackdb.
    # It is keyed by the name of the type of container and the value is
    # a dictionary of the container objects keyed by name with values as the
    # container objects. Note that individual Tracks are not included in this
    # dictionary; they are added directly to their configured parent container
    # (e.g., composite or trackdb or whatever).
    #
    # E.g.,
    #
    #   {
    #       'trackdb': TrackDB object,
    #       'composite': {
    #           'composite1': CompositeTrack object,
    #           'composite2': CompositeTrack object,
    #           ...
    #       },
    #       'hub': Hub object,
    #       'aggregate': {
    #           'agg1': AggregateTrack object,
    #       },
    #       ...
    #   }
    d = defaultdict(lambda: defaultdict(dict))

    if "genome" in sheets:
        d = make_assembly(wb, d)
    else:
        d = make_default_hub(wb, d)

    # The super tracks are made first because they are highest in the heirarchy
    d = make_super_tracks(wb, d)

    # Next make the composite tracks because they have the second highest heirarchy
    d = make_composite_tracks(wb, d)

    # View tracks can only exist in composite tracks, so they are added next
    d = make_view_tracks(wb, d)

    # Aggregate tracks can exist inside a super track or in the trackdb
    d = make_aggregate_tracks(wb, d)

    # Now that all the container tracks are made, the tracks can be added
    # We also keep track of which subgroups belong to which composite tracks
    d, comp2subgroup = add_tracks_and_make_subgroups(wb, d, sheets)

    # Add the subgroups to the respective composite tracks
    d = add_subgroups_to_composite(comp2subgroup, d)

    hub = d["hub"]

    # Create a directory with symlinked files and necessary text files with
    # directory structure
    trackhub.upload.stage_hub(hub=hub, staging=staging)


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument(
        "excel_file",
        help="Location of Excel file. See documentation for required formatting",
    )
    ap.add_argument(
        "--staging",
        help="Name of staging directory. Default is %(default)s",
        default="staging",
    )
    ap.add_argument("--verbose", help="verbose mode", action="store_true")
    args = ap.parse_args()

    main(filename=args.excel_file, staging=args.staging)
